# -*- coding: utf-8 -*-
"""P1 채점기: articles_clean.jsonl(정제 결과)을 골든셋(cleaned_articles_ex.xlsx)과 대조.

판정: 공백 표기 차이를 무시한(모든 공백 제거) 본문 완전 일치.
불일치 기사는 첫 분기점 주변을 양쪽으로 보여줘 규칙 개선 단서를 만든다.
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

from src import config
from src.p1_clean import PIPELINE_VERSION


def collapse(s: str) -> str:
    return "".join(ch for ch in s if not ch.isspace())


def load_golden(path: Path) -> dict[str, str]:
    wb = load_workbook(path, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(h).strip() for h in rows[0]]
    gi = {h: i for i, h in enumerate(header)}
    out = {}
    for r in rows[1:]:
        if r[gi["url"]] is None:
            continue
        text = str(r[gi["text"]] or "")
        if text.strip().lower() == "null":  # 골든셋의 '본문 없음' 표기 규약
            text = ""
        out[str(r[gi["article ID"]]).strip()] = text
    return out


def first_divergence(a: str, b: str) -> int:
    n = min(len(a), len(b))
    for i in range(n):
        if a[i] != b[i]:
            return i
    return n


def main(argv=None) -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    ap = argparse.ArgumentParser(description="P1 정제 결과를 골든셋과 대조 채점")
    ap.add_argument("--clean", type=Path, default=Path("data/articles_clean.jsonl"))
    ap.add_argument("--golden", type=Path,
                    default=config.part1_dir() / "cleaned_articles_ex.xlsx")
    ap.add_argument("--detail", type=int, default=8, help="불일치 상세 출력 건수")
    args = ap.parse_args(argv)

    golden = load_golden(args.golden)
    total = match = 0
    mismatches = []
    with open(args.clean, encoding="utf-8") as f:
        for line in f:
            a = json.loads(line)
            aid = a["article_id"]
            if aid not in golden:
                continue
            total += 1
            ours, gold = collapse(a["text"]), collapse(golden[aid])
            if ours == gold:
                match += 1
            else:
                mismatches.append((aid, a["title"], a["text"], golden[aid], ours, gold))

    print(f"[{PIPELINE_VERSION}] 골든셋 일치: {match}/{total} ({match / total * 100:.0f}%)")

    for aid, title, raw_ours, raw_gold, ours, gold in mismatches[:args.detail]:
        d = first_divergence(ours, gold)
        print(f"\n✗ {aid} | {title[:38]}")
        print(f"  길이(공백 제거): 정제 {len(ours)} vs 골든 {len(gold)} / 분기점 {d}")
        print(f"  정제: …{ours[max(0, d - 25):d]}▶{ours[d:d + 45]}…")
        print(f"  골든: …{gold[max(0, d - 25):d]}▶{gold[d:d + 45]}…")
    if len(mismatches) > args.detail:
        print(f"\n(그 외 불일치 {len(mismatches) - args.detail}건 생략)")


if __name__ == "__main__":
    main()
