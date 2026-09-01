# -*- coding: utf-8 -*-
"""P0 적재: 기사 원본(xlsx · csv · json/jsonl · 크롤링 dict) → 표준 기사 JSONL.

표준 기사 코어 스키마(크롤링 현실 기준): article_id · title · posted_date · url · text
- article_id는 URL 해시("A" + sha1[:8]) — 기사 추가·재정렬·재실행에도 불변
- 본문 text는 그대로 통과시킨다 — 정제는 P1 소관 (원본 보존 원칙)

**입력 형식은 이 파일의 read_* 함수만 안다.** transform 이후 파이프라인 전체는 형식을
모르며, 그래서 xlsx(선별 60건) → csv(조선일보 2,708건) → 크롤링 1건으로 입력이 바뀌어도
P1~P3는 손대지 않는다.

적재 정책 2종
    strict  선별본(xlsx)용 — 한 행이라도 검증 실패면 **전체 실패**(부분 산출 금지)
    bulk    대량 원본(csv)·크롤링용 — 불량 행을 articles_rejected.jsonl로 **격리**하고 계속
            (news.csv 실측: 중복 URL 10 · 작성일 결측 · 빈 본문 2 — 전부 아니면 전무를
             적용하면 2,708건이 통째로 죽는다. 격리도 기록이므로 전수 회계는 유지된다)

전수 회계 인바리언트: len(articles) + len(excluded) + len(rejected) == len(rows)

사용:
    python -m src.p0_load --input D:/part1/articles.xlsx --outdir data
    python -m src.p0_load --input D:/part1/news.csv --outdir data/bulk
    python -m src.p0_load --input crawled.json --outdir data/one
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, NamedTuple

from src import config

REQUIRED_COLUMNS = ["title", "posted date", "url", "text"]
VALID_CLASSES = (1, 2, 3, 4)
EXCLUDED_CLASS = 4  # 크롤링 오류(본문 하단 잘림) — 파이프라인 미투입 확정

AUX_FIELDNAMES = ["url", "article_id", "news_source", "query", "journalist",
                  "temp_class", "source_label"]

# 헤더 별칭 — 정본 키(왼쪽)로 통일한다. 소문자·공백 제거 후 비교하므로
# "posted date" / "posted_date" / "Posted Date"는 모두 같은 것으로 본다.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "title":                    ("title", "기사제목", "제목"),
    "posted date":              ("posted date", "posted_date", "작성일", "게시일", "date"),
    "url":                      ("url", "링크", "기사url", "link"),
    "text":                     ("text", "기사 본문 전체", "본문", "기사본문", "content", "body"),
    "temporary classification": ("temporary classification", "temp_class", "분류"),
    "article ID":               ("article id", "article_id", "기사id"),
    "news source":              ("news source", "news_source", "출처"),
    "query":                    ("query", "검색어"),
    "journalist":               ("journalist", "기자", "기자명"),
    "source_label":             ("검색 구분 레이블", "source_label", "label"),
    # 84차: 재크롤러(p_crawl) 산출의 구조 필드 — 있으면 통과, 없으면(구 데이터) 미기재
    "paragraphs":               ("paragraphs", "문단"),
    "publisher":                ("publisher", "언론사"),
    "subtitle":                 ("subtitle", "부제"),
}
CSV_EXT = {".csv", ".tsv"}
XLSX_EXT = {".xlsx", ".xlsm"}
JSON_EXT = {".json", ".jsonl", ".ndjson"}


class LoadResult(NamedTuple):
    articles: list[dict]
    aux: list[dict]
    excluded: list[dict]
    rejected: list[dict]


def make_article_id(url: str) -> str:
    """URL 기반 결정적 기사 ID. 앞뒤 공백·끝 슬래시 차이는 같은 기사로 본다."""
    normalized = url.strip().rstrip("/")
    return "A" + hashlib.sha1(normalized.encode("utf-8")).hexdigest()[:8]


def normalize_posted_date(value) -> str:
    """datetime/date/문자열 작성일을 'YYYY-MM-DD' 문자열로 통일.

    크롤링 원본에는 "2025.06.23. 14:52", "2025-06-23T09:00:00Z" 같은 표기가 섞이므로
    앞머리의 연-월-일만 취한다(시각은 버린다 — period 해소는 일 단위까지만 쓴다).
    """
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = "" if value is None else str(value).strip()
    m = re.match(r"^(\d{4})[-./년]\s*(\d{1,2})[-./월]\s*(\d{1,2})", s)
    if m:
        y, mo, d = (int(g) for g in m.groups())
        try:
            return date(y, mo, d).isoformat()
        except ValueError as e:
            raise ValueError(f"posted date 해석 불가: {value!r} ({e})") from None
    raise ValueError(f"posted date 해석 불가: {value!r}")


# ── 소스 어댑터 — 형식을 아는 유일한 층 ──────────────────────────────
def _canonical_header(header: Iterable) -> dict[str, int]:
    """실제 헤더 → {정본 키: 열 인덱스}. 별칭 해소 + 미지 컬럼은 원래 이름으로 통과."""
    idx: dict[str, int] = {}
    for i, raw in enumerate(header):
        name = "" if raw is None else str(raw).strip().lstrip("\ufeff")
        if not name:
            continue
        key = name.lower().replace(" ", "").replace("_", "")
        canon = next((c for c, alts in COLUMN_ALIASES.items()
                      if key in {a.lower().replace(" ", "").replace("_", "") for a in alts}),
                     name)
        idx.setdefault(canon, i)
    return idx


def _rows_from_table(header: Iterable, body: Iterable[Iterable]) -> list[dict]:
    idx = _canonical_header(header)
    missing = [c for c in REQUIRED_COLUMNS if c not in idx]
    if missing:
        raise ValueError(f"필수 컬럼 누락: {missing} / 실제 헤더: {list(header)}")
    rows = []
    for r in body:
        r = list(r)
        if not r or all(v is None or str(v).strip() == "" for v in r):
            continue
        rows.append({h: (r[i] if i < len(r) else None) for h, i in idx.items()})
    return rows


def read_rows_xlsx(path: Path) -> list[dict]:
    """xlsx 첫 시트를 헤더 기준 dict 행 목록으로 읽는다. 완전 빈 행은 무시."""
    from openpyxl import load_workbook          # csv 전용 실행에 xlsx 의존을 강제하지 않는다

    wb = load_workbook(path, read_only=True)
    try:
        rows_iter = wb.worksheets[0].iter_rows(values_only=True)
        return _rows_from_table(next(rows_iter), rows_iter)
    finally:
        wb.close()


def read_rows_csv(path: Path, encoding: str = "utf-8-sig") -> list[dict]:
    """csv/tsv를 읽는다. 기본 인코딩은 utf-8-sig(news.csv 실측: UTF-8 BOM).

    본문에 개행·따옴표가 그대로 들어 있어 csv.reader의 인용 처리를 반드시 거쳐야 한다.
    필드 크기 상한을 올리는 이유: 기사 본문 한 필드가 기본 상한(131,072자)을 넘길 수 있다.
    """
    try:
        csv.field_size_limit(sys.maxsize)
    except OverflowError:                        # Windows: C long 상한
        csv.field_size_limit(2**31 - 1)
    delim = "\t" if path.suffix.lower() == ".tsv" else ","
    for enc in (encoding, "cp949", "utf-8"):     # 인코딩 오판은 첫 줄에서 바로 드러난다
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f, delimiter=delim)
                return _rows_from_table(next(reader), reader)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"csv 인코딩 해석 실패: {path} (시도: {encoding}, cp949, utf-8)")


def read_rows_json(path: Path, encoding: str = "utf-8-sig") -> list[dict]:
    """json(객체 1건 또는 배열) · jsonl을 읽는다 — 실서비스의 '크롤링 기사 1건' 경로.

    기본 인코딩이 utf-8이 아니라 utf-8-sig인 이유: Windows 도구(PowerShell
    `Set-Content -Encoding UTF8` 등)가 BOM을 붙여 저장하는데, json 모듈은 BOM을
    만나면 곧바로 JSONDecodeError를 낸다. utf-8-sig는 BOM이 없는 파일도 정상 처리한다.
    """
    raw = path.read_text(encoding=encoding).strip()
    if not raw:
        return []
    if path.suffix.lower() in (".jsonl", ".ndjson"):
        records = [json.loads(ln) for ln in raw.splitlines() if ln.strip()]
    else:
        parsed = json.loads(raw)
        records = parsed if isinstance(parsed, list) else [parsed]
    return [_canonical_record(r) for r in records]


def _canonical_record(rec: dict) -> dict:
    """dict 1건의 키를 정본 키로 해소 — 크롤러 출력 키 이름 차이를 흡수한다."""
    if not isinstance(rec, dict):
        raise ValueError(f"기사 레코드는 객체여야 합니다: {type(rec).__name__}")
    idx = _canonical_header(rec.keys())
    keys = list(rec.keys())
    return {canon: rec[keys[i]] for canon, i in idx.items()}


def load_source(path: Path, fmt: str = "auto", encoding: str | None = None) -> list[dict]:
    """확장자(또는 --format)로 어댑터를 고른다. 반환은 형식 무관한 dict 행 목록."""
    path = Path(path)
    suffix = path.suffix.lower()
    if fmt == "auto":
        fmt = ("xlsx" if suffix in XLSX_EXT else
               "csv" if suffix in CSV_EXT else
               "json" if suffix in JSON_EXT else "")
        if not fmt:
            raise ValueError(f"입력 형식을 알 수 없습니다: {path.name} — --format 으로 지정하세요")
    if fmt == "xlsx":
        return read_rows_xlsx(path)
    if fmt == "csv":
        return read_rows_csv(path, encoding or "utf-8-sig")
    if fmt == "json":
        return read_rows_json(path, encoding or "utf-8-sig")
    raise ValueError(f"지원하지 않는 형식: {fmt!r}")


# 구 이름 — 기존 호출부·테스트 호환(xlsx 전용이던 시절의 이름)
read_rows = read_rows_xlsx


def default_policy(path: Path | None, fmt: str = "auto") -> str:
    """선별본(xlsx)은 strict, 대량 원본·크롤링(csv/json)은 bulk가 기본."""
    if path is None:
        return "bulk"
    suffix = Path(path).suffix.lower()
    if fmt == "xlsx" or (fmt == "auto" and suffix in XLSX_EXT):
        return "strict"
    return "bulk"


# ── 표준화 ────────────────────────────────────────────────────────────
def transform(rows: list[dict], policy: str = "strict",
              require_class: bool | None = None) -> LoadResult:
    """행 검증 + 표준화. 반환: LoadResult(articles, aux, excluded, rejected).

    policy="strict"  검증 실패가 하나라도 있으면 전체 실패(ValueError에 전 행 사유 나열)
    policy="bulk"    불량 행을 rejected로 격리하고 계속 — 기록이므로 회계는 유지된다
    require_class    temporary classification 필수 여부. 기본은 strict일 때만 필수
                     (크롤링 기사에는 이 라벨이 존재하지 않는다 — 1번의 작업용 라벨이다)
    """
    if policy not in ("strict", "bulk"):
        raise ValueError(f"policy는 strict|bulk 여야 합니다: {policy!r}")
    if require_class is None:
        require_class = policy == "strict"

    errors: list[str] = []
    articles: list[dict] = []
    aux: list[dict] = []
    excluded: list[dict] = []
    rejected: list[dict] = []
    seen_urls: dict[str, int] = {}
    seen_ids: dict[str, int] = {}

    def _reject(lineno: int, row: dict, reasons: list[str]) -> None:
        rejected.append({
            "lineno": lineno,
            "url": str(row.get("url") or "").strip(),
            "title": str(row.get("title") or "").strip()[:120],
            "reason_code": "INVALID_ROW",
            "reasons": reasons,
        })

    for lineno, row in enumerate(rows, start=2):  # 표 기준 행 번호(1행은 헤더)
        row_errors: list[str] = []

        url = str(row.get("url") or "").strip()
        # 제목의 BOM은 내용이 아니라 파일 아티팩트다(news.csv 실측: 중간 행에도 섞여 있음).
        # 본문은 건드리지 않는다 — 정제는 P1 소관(원본 보존).
        title = str(row.get("title") or "").strip().lstrip("﻿").strip()
        text_raw = row.get("text")
        text = "" if text_raw is None else str(text_raw)  # 원본 보존 — strip도 하지 않음

        if not url.lower().startswith("http"):
            row_errors.append(f"{lineno}행: url 비정상 {url!r}")
        if not title:
            row_errors.append(f"{lineno}행: title 비어 있음")
        paras_raw = row.get("paragraphs")
        if not text.strip() and not isinstance(paras_raw, list):
            # 크롤러 페이로드(paragraphs 키 보유)의 빈 본문은 결함이 아니라 실측이다
            # ([속보] 등 — 음성 대조군). 그 외 입력의 빈 text는 종전대로 반려(84차)
            row_errors.append(f"{lineno}행: text 비어 있음")
        if isinstance(paras_raw, list) and "\n".join(str(p) for p in paras_raw) != text:
            row_errors.append(f"{lineno}행: paragraphs와 text 불일치(크롤러 불변식 위반)")

        posted = None
        try:
            posted = normalize_posted_date(row.get("posted date"))
        except ValueError as e:
            row_errors.append(f"{lineno}행: {e}")

        temp_class = None
        tc_raw = row.get("temporary classification")
        has_class = tc_raw is not None and str(tc_raw).strip() != ""
        if has_class or require_class:
            try:
                temp_class = int(float(str(tc_raw).strip()))
                if temp_class not in VALID_CLASSES:
                    raise ValueError
            except (TypeError, ValueError):
                temp_class = None
                row_errors.append(f"{lineno}행: temporary classification 비정상 {tc_raw!r}")

        if url:
            if url in seen_urls:
                row_errors.append(f"{lineno}행: url 중복 (최초 {seen_urls[url]}행)")
            else:
                seen_urls[url] = lineno

        aid = make_article_id(url) if url else ""
        if not row_errors:
            provided_id = str(row.get("article ID") or "").strip()
            if provided_id and provided_id != aid:
                # 기재된 ID가 있으면 URL 계산값과 일치해야 한다 (수기 오류·복사 실수 방어)
                row_errors.append(
                    f"{lineno}행: article ID 불일치 — 기재 {provided_id!r} ≠ URL 계산 {aid!r}")
            elif aid in seen_ids:
                row_errors.append(f"{lineno}행: article_id 충돌 {aid} (최초 {seen_ids[aid]}행)")

        if row_errors:
            if policy == "strict":
                errors.extend(row_errors)
            else:
                _reject(lineno, row, row_errors)
            continue

        seen_ids[aid] = lineno
        aux.append({
            "url": url,
            "article_id": aid,
            "news_source": _s(row.get("news source")),
            "query": _s(row.get("query")),
            "journalist": _s(row.get("journalist")),
            "temp_class": temp_class,
            "source_label": _s(row.get("source_label")),
        })

        if temp_class == EXCLUDED_CLASS:
            excluded.append({
                "article_id": aid,
                "url": url,
                "title": title,
                "temp_class": temp_class,
                "reason_code": "CRAWL_ERROR_TRUNCATED",
                "reason": "크롤링 오류(본문 하단 잘림) — 원문 대조로 확인, 파이프라인 미투입",
            })
        else:
            art = {
                "article_id": aid,
                "title": title,
                "posted_date": posted,
                "url": url,
                "text": text,
            }
            # 84차: 재크롤 구조 필드 통과 — 코어 5필드 하위호환(구 데이터는 그대로 5필드).
            # paragraphs 정합(크롤러 불변식)은 row_errors 단계에서 이미 검증됨
            paras = row.get("paragraphs")
            if isinstance(paras, list) and paras:
                art["paragraphs"] = [str(p) for p in paras]
            for opt in ("publisher", "subtitle"):
                if _s(row.get(opt)):
                    art[opt] = _s(row.get(opt))
            articles.append(art)

    if errors:
        raise ValueError("P0 적재 검증 실패:\n" + "\n".join(errors))

    assert len(articles) + len(excluded) + len(rejected) == len(rows), "전수 회계 인바리언트 위반"
    return LoadResult(articles, aux, excluded, rejected)


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def standardize_article(record: dict) -> dict:
    """크롤링 기사 1건(dict) → 표준 기사 1건. 실서비스 진입점.

    키 이름은 별칭으로 흡수하고(제목/기사제목/title …), 검증 실패는 ValueError로 즉시 알린다
    — 1건 처리에서는 격리할 곳이 없으므로 호출자가 판단해야 한다.
    """
    result = transform([_canonical_record(record)], policy="strict", require_class=False)
    if not result.articles:
        raise ValueError("표준화 실패: 기사가 제외/거부되었습니다")
    return result.articles[0]


def write_outputs(outdir: Path, articles: list[dict], aux: list[dict],
                  excluded: list[dict], rejected: list[dict] | None = None) -> dict[str, Path]:
    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)
    paths = {
        "articles": outdir / "articles.jsonl",
        "aux": outdir / "aux_labels.csv",
        "excluded": outdir / "articles_excluded.jsonl",
    }
    with open(paths["articles"], "w", encoding="utf-8", newline="\n") as f:
        for a in articles:
            f.write(json.dumps(a, ensure_ascii=False) + "\n")
    with open(paths["excluded"], "w", encoding="utf-8", newline="\n") as f:
        for e in excluded:
            f.write(json.dumps(e, ensure_ascii=False) + "\n")
    if rejected:
        paths["rejected"] = outdir / "articles_rejected.jsonl"
        with open(paths["rejected"], "w", encoding="utf-8", newline="\n") as f:
            for r in rejected:
                f.write(json.dumps(r, ensure_ascii=False) + "\n")
    # 사이드카는 Excel에서 바로 열어보는 용도가 커서 BOM 포함(utf-8-sig)으로 쓴다
    with open(paths["aux"], "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=AUX_FIELDNAMES)
        w.writeheader()
        w.writerows(aux)
    return paths


def main(argv=None) -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    p = argparse.ArgumentParser(
        description="P0 적재: 기사 원본(xlsx·csv·json) → 표준 기사 JSONL + 사이드카 + 제외/거부 기록")
    p.add_argument("--input", type=Path, default=config.part1_dir() / "articles.xlsx",
                   help="입력 파일 (xlsx · csv/tsv · json/jsonl)")
    p.add_argument("--outdir", type=Path, default=config.data_dir(), help="산출물 디렉터리")
    p.add_argument("--format", choices=("auto", "xlsx", "csv", "json"), default="auto",
                   help="입력 형식 (기본 auto — 확장자로 판별)")
    p.add_argument("--policy", choices=("auto", "strict", "bulk"), default="auto",
                   help="auto: xlsx=strict, csv/json=bulk. strict는 불량 행 하나에 전체 실패")
    p.add_argument("--encoding", default=None, help="csv/json 인코딩 (기본 utf-8-sig → cp949 폴백)")
    p.add_argument("--limit", type=int, default=0, help="선두 N행만 적재 (대량 원본 시험용)")
    args = p.parse_args(argv)

    rows = load_source(args.input, args.format, args.encoding)
    if args.limit:
        rows = rows[:args.limit]
    policy = default_policy(args.input, args.format) if args.policy == "auto" else args.policy
    res = transform(rows, policy=policy)
    paths = write_outputs(args.outdir, res.articles, res.aux, res.excluded, res.rejected)

    dist: dict = {}
    for r in res.aux:
        dist[r["temp_class"]] = dist.get(r["temp_class"], 0) + 1
    print(f"입력 {len(rows)}건 = 적재 {len(res.articles)}건 + 제외 {len(res.excluded)}건 "
          f"+ 거부 {len(res.rejected)}건 — 전수 회계 OK (정책 {policy})")
    if dist:
        print(f"분류 분포: {dict(sorted(dist.items(), key=lambda kv: (kv[0] is None, kv[0])))}")
    for name, path in paths.items():
        print(f"  {name}: {path}")


if __name__ == "__main__":
    main()
