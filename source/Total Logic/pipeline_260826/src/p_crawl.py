# -*- coding: utf-8 -*-
"""P-1 크롤러 — 조선일보 기사 재크롤 + **문단 정보 확보** (CLAUDE.md 69차 설계 · 82차 구현).

왜 재크롤인가: 기존 news.csv 크롤은 개행이 전부 소실돼(P2 사전 스캔: 개행 0/52)
문단 번호를 억지로 부여할 수 없었고, 그것이 P2 최대 난제(무종결 소제목 경계 26건)의
근본 원인이었다. 상류에서 해소한다. 최종 서비스도 문단 구조의 본문이 필요하다
(기사 화면에 Claim 하이라이트 + 클릭 상세 — 문단 단위 렌더링).

추출 경로: 조선일보는 Arc(Fusion) 기반 — 본문 DOM은 클라이언트 렌더라 비어 있고,
서버 렌더된 `Fusion.globalContent` JSON의 `content_elements`(type=text)가 문단의 정본이다.
브라우저 불필요(plain HTTP + 브라우저 UA로 200 확인 — 82차 실측).

산출(JSONL 한 행 = 기사 하나):
  article_id · title · posted_date · url · text · paragraphs(+subtitle)
  - article_id: 입력 URL 기준 sha1(P0 규칙 불변 — 리다이렉트돼도 입력 URL로 계산)
  - text = "\n".join(paragraphs) — 하위호환(기존 파이프라인이 그대로 소비).
    기존 크롤과 달리 개행이 문단 경계 신호로 실린다(P2 개선 입력)
  - posted_date: display_date(UTC) → KST(+9h) 변환 후 날짜만 (조선일보 표기와 일치)

정책: 실패는 격리·기록(성공 + 실패 = 입력, §8-5 전수 회계) · 요청 간 딜레이(기본 1초).

사용:
  venv\\Scripts\\python.exe -m src.p_crawl --from-xlsx D:\\part1\\articles.xlsx --limit 60
  venv\\Scripts\\python.exe -m src.p_crawl --urls https://www.chosun.com/...
"""
from __future__ import annotations

import argparse
import html as _html
import json
import re
import time
import urllib.error
import urllib.request
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

from src import config
from src.p0_load import make_article_id

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")
KST = timezone(timedelta(hours=9))

# 언론사(브랜드) — 도메인 매핑이 정본(84차: 서비스 DB의 언론사 구분용). news.csv 전수
# 스캔에서 확인된 6개 도메인. 미지 도메인은 og:site_name → 도메인 문자열 순 폴백.
PUBLISHER_BY_DOMAIN = {
    "www.chosun.com": "조선일보",
    "news.chosun.com": "조선일보",
    "biz.chosun.com": "조선비즈",
    "economychosun.com": "이코노미조선",
    "health.chosun.com": "헬스조선",
    "realty.chosun.com": "땅집고",
    "weekly.chosun.com": "주간조선",
}
_RE_OG_SITE = re.compile(r'<meta[^>]*property="og:site_name"[^>]*content="([^"]*)"')


def detect_publisher(html_text: str, url: str) -> str:
    from urllib.parse import urlparse
    domain = urlparse(url).netloc
    if domain in PUBLISHER_BY_DOMAIN:
        return PUBLISHER_BY_DOMAIN[domain]
    m = _RE_OG_SITE.search(html_text)
    return _clean_text(m.group(1)) if m else domain

# Fusion 전역 컨텐츠 — 종료 앵커를 다음 Fusion 대입문으로 잡는다(본문 내 '};' 오절단 방지)
_RE_FUSION = re.compile(r"Fusion\.globalContent\s*=\s*(\{.*?\});\s*Fusion\.", re.S)
_RE_TAG = re.compile(r"<[^>]+>")


@dataclass
class CrawlResult:
    article_id: str
    title: str
    posted_date: str
    url: str
    text: str
    paragraphs: list[str] = field(default_factory=list)
    subtitle: str = ""            # 부제(subheadlines) — 본문 아님, 참고 보존
    publisher: str = ""           # 언론사 브랜드(84차 — 서비스 DB 구분용): 도메인 매핑 정본
    last_updated: str = ""        # 기사 수정 시각(ISO) — 리니지용
    extractor: str = ""           # 어느 파서가 뽑았나(99차 감사용): fusion·ndsoft·generic …

    def to_dict(self) -> dict:
        return asdict(self)


def fetch_html(url: str, timeout: int = 30, retries: int = 3) -> str:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    delay = 2.0
    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read().decode("utf-8", "replace")
        except urllib.error.HTTPError as e:
            if e.code in (429, 500, 502, 503) and attempt < retries:
                time.sleep(delay)
                delay *= 2
                continue
            raise
        except OSError:
            if attempt < retries:
                time.sleep(delay)
                delay *= 2
                continue
            raise


def _clean_text(s: str) -> str:
    """태그 제거 + 엔티티 해제 + 공백 정돈 — 문단 내부는 무수정 원칙에 가깝게 최소 처리."""
    s = _RE_TAG.sub("", s)
    s = _html.unescape(s)
    return re.sub(r"[ \t\u00a0]+", " ", s).strip()


def _kst_date(iso: str | None) -> str:
    if not iso:
        return ""
    try:
        dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
    except ValueError:
        return ""
    return dt.astimezone(KST).date().isoformat()


def parse_fusion(html_text: str, url: str) -> CrawlResult:
    """Fusion.globalContent JSON → CrawlResult (chosun.com·biz.chosun.com — Arc 계열).

    본문 문단 0개는 실패가 아니라 **빈 본문 기사**([속보] 등 — 기존 크롤은 노이즈
    746자를 본문인 척 실었던 유형)다. paragraphs=[]로 정직하게 내보낸다(음성 대조군).
    """
    m = _RE_FUSION.search(html_text)
    if not m:
        raise ValueError("Fusion.globalContent 블록 없음 — 페이지 구조 변경 또는 비기사 페이지")
    gc = json.loads(m.group(1))
    paragraphs = []
    for e in gc.get("content_elements", []):
        if e.get("type") == "text":
            if t := _clean_text(e.get("content", "")):
                paragraphs.append(t)
        elif e.get("type") == "quote":
            # 서문·편집자주가 blockquote로 실리는 기사 실재(Aac48a075 — 골든 정제본이
            # 본문으로 유지) → 내부 text들을 한 문단으로 합쳐 포함
            inner = [t for c in e.get("content_elements", [])
                     if c.get("type") == "text" and (t := _clean_text(c.get("content", "")))]
            if inner:
                paragraphs.append(" ".join(inner))
        # 그 외 타입은 **화이트리스트 방식으로 자동 제외** — 실측 확인(83차):
        #   interstitial_link('칼럼 전문 링크' 링크 위젯) · raw_html(기자 소개·프로모 박스)
        #   · image(캡션 포함). 구 크롤 텍스트에 섞여 들어가던 비본문이 원천 차단된다.
    title = _clean_text(gc.get("headlines", {}).get("basic", ""))
    subtitle = _clean_text(gc.get("subheadlines", {}).get("basic", ""))
    posted = (_kst_date(gc.get("display_date"))
              or _kst_date(gc.get("first_publish_date"))
              or _kst_date(gc.get("created_date")))
    return CrawlResult(
        article_id=make_article_id(url), title=title, posted_date=posted,
        url=url, text="\n".join(paragraphs), paragraphs=paragraphs,
        subtitle=subtitle, last_updated=gc.get("last_updated_date") or "",
    )


# 이코노미조선(economychosun.com) — 구형 CMS(html_dir 경로). 본문은
# <div id="articleBody"> 안의 <p> 태그들, 캡션(<figure>)은 <p> 밖이라 자연 배제.
_RE_ECON_P = re.compile(r"<p[^>]*>(.*?)</p>", re.S)
_RE_OG_TITLE = re.compile(r'<meta\s+property="og:title"\s+content="([^"]*)"')
_RE_ECON_DATE = re.compile(r"/html_dir/(\d{4})/(\d{2})/(\d{2})/")


def parse_economychosun(html_text: str, url: str) -> CrawlResult:
    i = html_text.find('id="articleBody"')
    if i < 0:
        raise ValueError("articleBody 블록 없음 — 이코노미조선 구조 변경")
    ends = [j for marker in ("기사본문 : e", "articleFooter", "기사하단")
            if (j := html_text.find(marker, i)) > 0]
    body = html_text[i:min(ends) if ends else len(html_text)]
    # 순수 식별자 토큰(예: 'plus_point_0' — 템플릿 아티팩트)은 본문이 아니다
    paragraphs = [t for p in _RE_ECON_P.findall(body)
                  if (t := _clean_text(p)) and not re.fullmatch(r"[A-Za-z0-9_\-]+", t)]
    tm = _RE_OG_TITLE.search(html_text)
    dm = _RE_ECON_DATE.search(url)
    return CrawlResult(
        article_id=make_article_id(url),
        title=_clean_text(tm.group(1)) if tm else "",
        posted_date="-".join(dm.groups()) if dm else "",
        url=url, text="\n".join(paragraphs), paragraphs=paragraphs,
    )


# 구형 조선 서브도메인(health·news·realty.chosun.com — html_dir 경로) 공용:
# 본문 = <div id="news_body_id"> 안, 문단 구분은 <br> 연속. 부제는 <h3 class="news_subtitle">.
_RE_BR = re.compile(r"<br\s*/?>", re.I)
_RE_LEGACY_SUB = re.compile(r'<h3 class="news_subtitle">(.*?)</h3>', re.S)
_RE_LEGACY_NOISE = re.compile(
    r"^\|?\s*(입력|수정)\s*:?\s*\d{4}|^Copyright|조선일보\s*&|무단 전재|재배포 금지|^관련\s*기사")
_RE_LEGACY_IMGBOX_OPEN = re.compile(r'<div[^>]*class="[^"]*img[^"]*"[^>]*>')
_RE_DIV_TOKEN = re.compile(r"<div\b|</div>")


def _cut_img_blocks(s: str) -> str:
    """이미지·캡션 블록(class에 img가 든 div) 제거 — 중첩 div는 정규식으로 못 자르므로
    여닫이 깊이를 세는 균형 스캔으로 절단(소스 포매팅 무관 결정적)."""
    while m := _RE_LEGACY_IMGBOX_OPEN.search(s):
        depth, end = 1, len(s)
        for t in _RE_DIV_TOKEN.finditer(s, m.end()):
            depth += 1 if t.group(0) == "<div" else -1
            if depth == 0:
                end = t.end()
                break
        s = s[:m.start()] + s[end:]
    return s


def parse_chosun_legacy(html_text: str, url: str) -> CrawlResult:
    i = html_text.find('id="news_body_id"')
    if i < 0:
        raise ValueError("news_body_id 블록 없음 — 구형 조선 페이지 구조 변경")
    i = html_text.index(">", i) + 1          # 여는 태그 안에서 슬라이스하면 태그 조각이 샌다
    ends = [j for marker in ("기사 본문 end", "news_copyright", "copyright")
            if (j := html_text.find(marker, i)) > 0]
    body = html_text[i:min(ends) if ends else len(html_text)]
    sub = _RE_LEGACY_SUB.search(body)
    subtitle = _clean_text(_RE_BR.sub(" ", sub.group(1))) if sub else ""
    body = _RE_LEGACY_SUB.sub("", body)
    body = re.sub(r"<script.*?</script>|<style.*?</style>", "", body, flags=re.S)
    body = re.sub(r"<!--.*?(-->|$)", "", body, flags=re.S)   # 주석(말미 미종결 포함)
    body = _cut_img_blocks(body)
    # <br>·블록 태그 닫힘 = 문단/줄 경계 — 소스 포매팅(개행 유무)에 의존하지 않는다
    body = re.sub(r"</(p|div|h\d|li|tr)>", "\n", body, flags=re.I)
    lines = _RE_TAG.sub("", _RE_BR.sub("\n", body)).split("\n")
    paragraphs = [t for ln in lines
                  if (t := _clean_text(ln)) and not _RE_LEGACY_NOISE.search(t)]
    tm = _RE_OG_TITLE.search(html_text)
    dm = _RE_ECON_DATE.search(url)
    return CrawlResult(
        article_id=make_article_id(url),
        title=_clean_text(tm.group(1)) if tm else "",
        posted_date="-".join(dm.groups()) if dm else "",
        url=url, text="\n".join(paragraphs), paragraphs=paragraphs,
        subtitle=subtitle,
    )


# ndsoft 계열 CMS(weekly.chosun.com 등 — articleView.html?idxno= 경로):
# 본문 = <article itemprop="articleBody"> 안의 <p>, 사진·캡션은 <figure> 블록.
_RE_NDSOFT_BODY = re.compile(r'<article[^>]*itemprop="articleBody"[^>]*>', re.I)
_RE_NDSOFT_DATE = re.compile(
    r'<meta[^>]*property="article:published_time"[^>]*content="([^"]+)"')


def parse_ndsoft(html_text: str, url: str) -> CrawlResult:
    m = _RE_NDSOFT_BODY.search(html_text)
    if not m:
        raise ValueError("articleBody(ndsoft) 블록 없음 — 페이지 구조 변경")
    body = html_text[m.end():html_text.find("</article>", m.end())]
    body = re.sub(r"<figure.*?</figure>", "", body, flags=re.S)   # 사진·캡션
    body = re.sub(r"<script.*?</script>|<!--.*?(-->|$)", "", body, flags=re.S)
    # <p>가 문단 정본. 없으면(구형 편집) <br>·블록 태그 경계로 폴백
    ps = [t for pfrag in _RE_ECON_P.findall(body) if (t := _clean_text(pfrag))]
    if not ps:
        lines = _RE_TAG.sub("", re.sub(r"<br\s*/?>|</(p|div|h\d)>", "\n", body, flags=re.I))
        ps = [t for ln in lines.split("\n") if (t := _clean_text(ln))]
    tm = _RE_OG_TITLE.search(html_text)
    dm = _RE_NDSOFT_DATE.search(html_text)
    posted = ""
    if dm:
        try:
            posted = datetime.fromisoformat(dm.group(1)).astimezone(KST).date().isoformat()
        except ValueError:
            pass
    return CrawlResult(
        article_id=make_article_id(url),
        title=_clean_text(tm.group(1)) if tm else "",
        posted_date=posted, url=url,
        text="\n".join(ps), paragraphs=ps,
    )


# ── 범용 추출기(A안 · 99차) — 전용 파서가 없는 미지 도메인 폴백 ────────────────────
# 원리: 한국 언론 CMS들이 공통으로 쓰는 **본문 컨테이너 이름 관행**(articleBody ·
# article-body · news_body · entry-content …)을 id/class 이름으로 찾고, 그 안의
# <p>(없으면 <br>·블록 태그 닫힘)를 문단으로 삼는다. 외부 라이브러리 0 · 순수 정규식.
#
# 위치: 디스패처의 **마지막 폴백**이다. 전용 파서가 있는 도메인은 그쪽이 먼저 받으므로,
# 여기까지 오는 것은 "구조를 모르는 새 도메인"뿐이다. 여기서도 실패하면 종전대로
# 시끄럽게 실패한다(crawl_errors.jsonl) — 조용한 잘림은 여전히 없다.
#
# 한계: 서버가 본문을 안 실어 보내는 **JS 렌더 사이트는 원리적으로 못 잡는다**
#       (조선 Arc/Fusion 계열은 전용 파서가 서버 렌더 JSON을 읽어 해결한 것).
#       그런 도메인은 전용 파서를 추가하거나 헤드리스 렌더링 경로가 필요하다.

_GENERIC_HINTS = (
    "articlebody", "articlecontent", "articleview", "articletext", "articletxt",
    "newsbody", "newsct", "newsview", "newstext", "arttxt", "artbody",
    "entrycontent", "postcontent", "storybody", "contbody", "contentbody",
)
_RE_BLOCK_OPEN = re.compile(r"<(div|article|section)\b([^>]*)>", re.I)
# form을 넣지 않는다: ASP.NET 페이지(오마이뉴스 at_pg.aspx 등)는 **본문을 포함한 문서
# 전체를 <form> 하나로 감싼다** — 통째로 지우면 본문이 사라진다(실측: 추출 0자).
# form 안의 입력 요소는 어차피 텍스트를 남기지 않으므로 제거할 필요도 없다.
_RE_DROP_BLOCK = re.compile(
    r"<(script|style|noscript|iframe|figure|aside|nav|footer|svg)\b.*?</\1\s*>",
    re.S | re.I)
_RE_GENERIC_NOISE = re.compile(
    r"^[\s|\[\]·]*(입력|수정|등록|승인)\s*:?\s*\d{4}"          # 타임스탬프 줄
    r"|무단\s*전재|재배포\s*금지|저작권자|Copyright|ⓒ|©"
    r"|^[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}$"               # 이메일만 있는 줄
    r"|^관련\s*기사|^\s*\[[^\]]{1,20}\]\s*$"          # 태그·구분 줄
    r"|^\s*(광고|AD|ADVERTISEMENT)\s*$")               # 광고 라벨 줄
_RE_H1 = re.compile(r"<h1[^>]*>(.*?)</h1>", re.S | re.I)
_RE_TITLE_TAG = re.compile(r"<title[^>]*>(.*?)</title>", re.S | re.I)
# <meta> 파싱은 속성 순서·따옴표 종류에 무관해야 한다. 특히 구분자를 **역참조**로 잡는
# 이유: content 값 안에 반대쪽 따옴표가 든 기사가 실재한다(주간조선 실측 — 제목의
# 작은따옴표에서 [^"']* 방식이 잘려 나가 <title> 폴백으로 떨어졌다).
_RE_META_TAG = re.compile(r"<meta\b[^>]*>", re.I)
_RE_META_KEY = re.compile(r"""(?:property|name|itemprop)\s*=\s*(["'])(.*?)\1""", re.I)
_RE_META_CONTENT = re.compile(r"""content\s*=\s*(["'])(.*?)\1""", re.I | re.S)
_RE_JSONLD_DATE = re.compile(r'"datePublished"\s*:\s*"([^"]+)"')
_RE_TEXT_DATE = re.compile(r"(20\d{2})[.\-/년]\s?(\d{1,2})[.\-/월]\s?(\d{1,2})")
_RE_URL_DATE_ANY = re.compile(r"/(20\d{2})[/.\-]?(0[1-9]|1[0-2])[/.\-]?(0[1-9]|[12]\d|3[01])")


def _slice_balanced(s: str, tag: str, start: int) -> str:
    """여는 태그 뒤(start)부터 **같은 태그의 짝**이 닫히는 곳까지 — 중첩 안전.
    정규식으로는 표현할 수 없는 영역이라 여닫이 깊이를 직접 센다."""
    depth = 1
    for t in re.finditer(rf"<{tag}\b|</{tag}\s*>", s[start:], re.I):
        depth += -1 if t.group(0).startswith("</") else 1
        if depth == 0:
            return s[start:start + t.start()]
    return s[start:]


# 본문 컨테이너 안에 섞여 들어오는 비본문 블록 — 이름으로 걷어낸다(추천기사·댓글·
# 기자 프로필·광고·SNS). 미지 도메인이라 위치는 모르지만 이름 관행은 공통이다.
_GENERIC_JUNK = (
    "relat", "relart", "recommend", "reply", "comment", "banner", "advert", "googlead",
    "share", "sns", "copyright", "reporter", "byline", "promo", "popular",
    "ranking", "widget", "newsletter", "subscribe", "tagbox", "morenews",
)
_RE_JUNK_OPEN = re.compile(r"<(div|section|ul|aside)\b([^>]*)>", re.I)


def _drop_named_blocks(body: str) -> str:
    """이름이 잡스러운 블록을 통째로 제거 — 여는 태그부터 짝 닫는 태그까지.

    크기 가드: 컨테이너 텍스트의 40% 이상을 지우는 제거는 **하지 않는다.** 잡블록은
    본래 작고, 큰 삭제는 HTML의 짝이 안 맞아 균형 스캔이 본문까지 삼킨 신호다
    (주간조선 실측: sticky 앵커 div 하나를 지우자 본문 530자가 71자로 무너졌다).
    """
    base = len(_clean_text(body))
    while True:
        for m in _RE_JUNK_OPEN.finditer(body):
            attrs = re.sub(r"[-_\s]", "", m.group(2).lower())
            if not any(j in attrs for j in _GENERIC_JUNK):
                continue
            inner = _slice_balanced(body, m.group(1), m.end())
            if base and len(_clean_text(inner)) >= base * 0.4:
                continue
            pos = m.end() + len(inner)
            close = re.match(rf"</{m.group(1)}\s*>", body[pos:], re.I)
            body = body[:m.start()] + body[pos + (close.end() if close else 0):]
            break
        else:
            return body


_RE_A_TEXT = re.compile(r"<a\b[^>]*>(.*?)</a>", re.S | re.I)


def _mostly_link(frag: str) -> bool:
    """텍스트가 사실상 전부 링크인 조각 = 추천기사·목록. 본문 문단은 그렇지 않다.
    긴 조각은 예외로 둔다(본문 전체가 링크인 경우는 없다시피 하고, 오판 대가가 크다)."""
    txt = _clean_text(frag)
    if not txt or len(txt) > 200:
        return False
    return sum(len(_clean_text(a)) for a in _RE_A_TEXT.findall(frag)) >= len(txt) * 0.8


def _paragraphs_from_html(body: str) -> list[str]:
    """<p> 분리와 <br>·블록 태그 분리를 **둘 다 해 보고 텍스트가 많은 쪽**을 쓴다.
    구형 CMS는 빈 <p> 하나만 두고 본문은 <br>로 흘리는 경우가 있어(realty 실측)
    '<p>가 있으면 <p>' 규칙 하나로는 본문을 통째로 놓친다."""
    body = _drop_named_blocks(body)
    by_p = [t for frag in _RE_ECON_P.findall(body)
            if not _mostly_link(frag) and (t := _clean_text(frag))]
    broken = re.sub(r"<br\s*/?>|</(p|div|h\d|li|blockquote|tr)\s*>", "\n", body, flags=re.I)
    by_br = [t for ln in broken.split("\n")
             if not _mostly_link(ln) and (t := _clean_text(ln))]
    keep = lambda ps: [p for p in ps if len(p) > 1 and not _RE_GENERIC_NOISE.search(p)]
    a, b = keep(by_p), keep(by_br)
    return a if sum(map(len, a)) >= sum(map(len, b)) else b


def _meta_value(html_text: str, keys: tuple[str, ...]) -> str:
    """<meta>에서 키(property·name·itemprop)에 맞는 content — 속성 순서 무관."""
    for tag in _RE_META_TAG.findall(html_text):
        k, c = _RE_META_KEY.search(tag), _RE_META_CONTENT.search(tag)
        if k and c and k.group(2).strip().lower() in keys:
            return c.group(2)
    return ""


def _generic_title(html_text: str) -> str:
    if t := _clean_text(_meta_value(html_text, ("og:title", "title"))):
        return t
    for rx in (_RE_H1, _RE_TITLE_TAG):
        if (m := rx.search(html_text)) and (t := _clean_text(m.group(1))):
            return t
    return ""


def _generic_date(html_text: str, url: str) -> str:
    """**URL 경로 날짜 우선**, 없으면 메타 → JSON-LD.

    경로 날짜를 먼저 보는 이유: ① 기존 전용 파서(chosun_legacy·economychosun)가
    같은 규칙이라 코퍼스 안에서 일관되고 ② 한국 언론이 KST 시각에 Z(UTC)를 붙여
    내보내는 경우가 있어 메타를 변환하면 날짜가 하루 밀린다(health 실측: 경로
    10-10 ↔ 메타 변환 10-11). 타임존이 없는 표기는 현지 시각으로 간주한다
    (astimezone에 맡기면 실행 머신 로캘에 따라 날짜가 달라진다 — 재현성 우선).
    """
    if m := _RE_URL_DATE_ANY.search(url):
        return "-".join(m.groups())
    raws = [_meta_value(html_text, ("article:published_time", "datepublished",
                                    "pubdate", "og:regdate", "date"))]
    if m := _RE_JSONLD_DATE.search(html_text):
        raws.append(m.group(1))
    for raw in [r.strip() for r in raws if r and r.strip()]:
        try:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
            return (dt.date() if dt.tzinfo is None else dt.astimezone(KST).date()).isoformat()
        except ValueError:
            if d := _RE_TEXT_DATE.search(raw):
                return f"{d.group(1)}-{int(d.group(2)):02d}-{int(d.group(3)):02d}"
    return ""


def parse_generic(html_text: str, url: str, min_chars: int = 100) -> CrawlResult:
    """이름 관행 기반 본문 추출 — 미지 도메인용 폴백(99차).

    본문 컨테이너를 id/class 이름으로 찾고(§_GENERIC_HINTS), 잡블록을 이름으로
    걷어낸 뒤 문단을 뽑는다. 본문이 너무 적으면 **성공을 가장하지 않고 실패**한다
    — 절반만 뽑아 놓고 정상인 척하는 것이 이 파이프라인에서 가장 나쁜 실패다.
    """
    src = re.sub(r"<!--.*?(-->|$)", " ", _RE_DROP_BLOCK.sub(" ", html_text), flags=re.S)
    cands: list[tuple[int, int, str, bool]] = []
    for m in _RE_BLOCK_OPEN.finditer(src):
        attrs = re.sub(r"[-_\s]", "", m.group(2).lower())
        if any(h in attrs for h in _GENERIC_HINTS):
            inner = _slice_balanced(src, m.group(1), m.end())
            cands.append((m.end(), m.end() + len(inner), inner,
                          'itemprop="articlebody"' in attrs or "itemprop='articlebody'" in attrs))
    if not cands and (m := re.search(r"<article\b[^>]*>", src, re.I)):
        inner = _slice_balanced(src, "article", m.end())   # 시맨틱 태그 최후 폴백
        cands.append((m.end(), m.end() + len(inner), inner, False))

    # 선택 규칙
    #  ① itemprop="articleBody"(schema.org)가 있으면 **그것만** 본다 — 이름 추측보다
    #     발행사가 명시한 시맨틱 표시가 우선이다(ndsoft 계열 CMS가 이 표기를 쓴다).
    #  ② 없으면 가장 큰 후보의 70% 이상을 담은 것들 중 가장 좁은 것.
    #     - 가장 큰 것만 쓰면 '본문 + 추천기사' 바깥 래퍼가 이기고,
    #     - 가장 안쪽 것만 쓰면 본문을 여러 조각에 나눠 담은 CMS에서 한 조각만 잡힌다
    #       (이코노미조선 실측: article--contents-text 3조각 중 하나 = 본문 1/3).
    paragraphs: list[str] = []
    if cands:
        semantic = [c for c in cands if c[3]]
        pool = semantic or cands
        if not semantic:
            sizes = {c[0]: len(_clean_text(c[2])) for c in pool}
            cut = max(sizes.values()) * 0.7
            pool = [c for c in pool if sizes[c[0]] >= cut]
        paragraphs = _paragraphs_from_html(min(pool, key=lambda c: c[1] - c[0])[2])
    total = sum(map(len, paragraphs))
    if total < min_chars:
        raise ValueError(f"범용 추출 실패 — 본문 컨테이너 미탐지 또는 본문 과소({total}자). "
                         f"JS 렌더 페이지이거나 새 구조 — 전용 파서 필요")
    return CrawlResult(
        article_id=make_article_id(url), title=_generic_title(html_text),
        posted_date=_generic_date(html_text, url), url=url,
        text="\n".join(paragraphs), paragraphs=paragraphs,
    )


def parse_article(html_text: str, url: str) -> CrawlResult:
    """도메인 디스패처 — 전용 파서 4종 → **범용 추출기 폴백** → 시끄러운 실패.

    ① 이코노미조선 ② Arc(Fusion — www·biz) ③ 구형 서브도메인(news_body_id)
    ④ ndsoft 계열(itemprop=articleBody) ⑤ **범용(이름 관행 기반, 99차)**

    ⑤ 덕분에 처음 보는 도메인도 서버가 본문을 HTML에 실어 보내기만 하면 받아낸다.
    그래도 실패하면 예외 → crawl_errors.jsonl (조용한 잘림은 여전히 없다).
    """
    if "economychosun.com" in url:
        r, how = parse_economychosun(html_text, url), "economychosun"
    elif _RE_FUSION.search(html_text):
        r, how = parse_fusion(html_text, url), "fusion"
    elif 'id="news_body_id"' in html_text:
        r, how = parse_chosun_legacy(html_text, url), "chosun_legacy"
    elif _RE_NDSOFT_BODY.search(html_text):
        r, how = parse_ndsoft(html_text, url), "ndsoft"   # 주간조선 등(90차)
    else:
        r, how = parse_generic(html_text, url), "generic"  # 미지 도메인(99차)
    r.publisher = detect_publisher(html_text, url)
    r.extractor = how
    return r


def crawl_urls(urls: list[str], out_path: Path, err_path: Path,
               delay: float = 1.0, resume: bool = False) -> dict:
    """URL 목록 재크롤 — 성공은 out, 실패는 err에 기록(전수 회계 3항: 성공+실패=입력).

    resume=True: out 파일에 이미 있는 URL은 건너뛰고 **이어서 append** — 전량(2,701건
    ~45분) 실행이 중간에 끊겨도 처음부터 다시 돌지 않는다. 실패했던 URL은 재시도한다.
    """
    # 완료 대조는 URL 문자열이 아니라 article_id(정규화 해시) — 입력원마다 말미 슬래시
    # 등 표기가 달라도 같은 기사면 건너뛴다
    done: set[str] = set()
    if resume and out_path.exists():
        with open(out_path, encoding="utf-8") as f:
            done = {json.loads(line)["article_id"] for line in f if line.strip()}
    todo = [u for u in urls if make_article_id(u) not in done]
    if done:
        print(f"resume: 완료 {len(done)}건 보유 · 입력 {len(urls)}건 중 "
              f"{len(urls) - len(todo)}건 건너뜀 → 남은 {len(todo)}건")

    ok, failed = [], []
    out_path.parent.mkdir(parents=True, exist_ok=True)
    mode = "a" if (resume and out_path.exists()) else "w"
    with open(out_path, mode, encoding="utf-8") as f:
        for i, url in enumerate(todo):
            if i:
                time.sleep(delay)
            try:
                r = parse_article(fetch_html(url), url)
                ok.append(r)
                f.write(json.dumps(r.to_dict(), ensure_ascii=False) + "\n")
                f.flush()                            # 끊겨도 지금까지의 성공분은 보존
                n = len(r.paragraphs)
                print(f"[{i + 1}/{len(todo)}] OK  {make_article_id(url)} 문단 {n}개"
                      + (" ⚠ 빈 본문" if n == 0 else ""))
            except Exception as e:                   # 실패 격리 — 한 건이 전체를 죽이지 않는다
                failed.append({"url": url, "article_id": make_article_id(url),
                               "error": f"{type(e).__name__}: {e}"[:200]})
                print(f"[{i + 1}/{len(todo)}] FAIL {make_article_id(url)} {failed[-1]['error'][:80]}")
    with open(err_path, "w", encoding="utf-8") as f:
        for row in failed:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")
    assert len(ok) + len(failed) == len(todo), "전수 회계 위반"
    return {"input": len(urls), "skipped_done": len(done) and len(urls) - len(todo),
            "ok": len(ok), "failed": len(failed),
            "out": str(out_path), "errors": str(err_path)}


def _urls_from_xlsx(path: Path) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr = [str(h).strip().lower() if h else "" for h in rows[0]]
    ui = hdr.index("url")
    return [str(r[ui]).strip() for r in rows[1:] if r[ui]]


def _urls_from_csv(path: Path) -> list[str]:
    """news.csv 등에서 URL 열 추출 — http로 시작하는 값만(비정상 행은 애초에 제외)."""
    import csv
    csv.field_size_limit(10**9)                      # 본문 열이 13만 자를 넘는 행 실재
    with open(path, encoding="utf-8-sig", newline="") as f:
        rdr = csv.DictReader(f)
        ucol = next(c for c in rdr.fieldnames if "url" in c.lower())
        return [u for row in rdr
                if (u := (row.get(ucol) or "").strip()).lower().startswith("http")]


def main() -> None:
    ap = argparse.ArgumentParser(description="조선일보 재크롤 — 문단 정보 포함")
    ap.add_argument("--from-xlsx", type=Path, help="url 열을 가진 xlsx (articles.xlsx)")
    ap.add_argument("--from-csv", type=Path, help="URL 열을 가진 csv (news.csv — 전량 재크롤)")
    ap.add_argument("--urls", nargs="*", default=[], help="직접 지정 URL")
    ap.add_argument("--resume", action="store_true",
                    help="out 파일에 이미 있는 URL은 건너뛰고 이어서 수집(중단 재개용)")
    ap.add_argument("--out", type=Path, default=config.data_dir() / "articles_crawled.jsonl")
    ap.add_argument("--errors", type=Path, default=config.data_dir() / "crawl_errors.jsonl")
    ap.add_argument("--delay", type=float, default=1.0, help="요청 간 초 (기본 1.0 — 예의)")
    ap.add_argument("--limit", type=int, default=0)
    args = ap.parse_args()

    urls = list(args.urls)
    if args.from_xlsx:
        urls += _urls_from_xlsx(args.from_xlsx)
    if args.from_csv:
        urls += _urls_from_csv(args.from_csv)
    # 중복 제거(순서 유지) — 같은 기사 재요청 방지(news.csv 실측: 중복 URL 10종)
    urls = list(dict.fromkeys(urls))
    if args.limit:
        urls = urls[:args.limit]
    if not urls:
        ap.error("--from-xlsx / --from-csv / --urls 중 하나 필요")
    summary = crawl_urls(urls, args.out, args.errors, delay=args.delay, resume=args.resume)
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
