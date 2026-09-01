# -*- coding: utf-8 -*-
"""골든셋(claim_silver_set_ver3.xlsx) 로더 — DocumentSet으로 변환 + 버전 해시.

골든 xlsx 스키마(17열): claim_id · article_id · sent_id · posted_date · claim · metric ·
metric_normalized · value · unit · value_type · direction · period · comparison_basis ·
forecast · kosis_eligible · exclusion_code · note
- claim_id 있는 행 = Claim / 없는 행(exclusion_code만) = 제외 대장
- kosis_eligible 셀은 "TRUE"/"FALSE" 문자열 — 로드 시 bool로, §4.8 파생식과 일치 검증 가능
"""
from __future__ import annotations

import hashlib
from pathlib import Path

from openpyxl import load_workbook

from src import config
from src.p3_schemas import ClaimRecord, ExcludedRecord, DocumentSet

# 골든셋은 저장소에 포함되지 않는 자료(수작업 저작본) — 위치는 .env의 PART1_DIR로 지정한다.
# 89차: 새 크롤 체인으로 이관된 ver3이 정본(구 ver2 + 회복 8기사 부록 병합).
GOLDEN_DEFAULT = config.part1_dir() / "claim_silver_set_ver3.xlsx"

# 79차 열 재배치: comparison_period 신설(comparison_basis 뒤) ·
# metric_normalized는 당장 안 쓰므로 맨 뒤로(사용자 지시)
_COLS = [
    "claim_id", "article_id", "sent_id", "posted_date", "claim",
    "metric", "value", "unit", "value_type", "direction",
    "period", "comparison_basis", "comparison_period",
    "forecast", "kosis_eligible", "exclusion_code", "note", "metric_normalized",
]


def _s(v) -> str:
    """셀 값 → 트리밍 문자열(None → ''). 날짜·숫자·불리언 셀 방어 파싱(§6 채점기 규약).

    - datetime/date: Excel이 period·posted_date를 날짜로 자동 변환한 경우 → ISO 날짜 문자열
      (자정 시각이면 'YYYY-MM-DD', 그 외는 isoformat 그대로 남겨 형식 검사에 걸리게 둔다)
    - float 정수값: '86.0' 오염 방지 → '86'
    - bool: 'TRUE'/'FALSE'
    """
    import datetime as _dt
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, _dt.datetime):
        return v.date().isoformat() if (v.hour, v.minute, v.second) == (0, 0, 0) else v.isoformat()
    if isinstance(v, _dt.date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def golden_version(path: Path) -> str:
    return hashlib.sha1(Path(path).read_bytes()).hexdigest()[:8]


def load_golden(path: Path | str = GOLDEN_DEFAULT) -> DocumentSet:
    path = Path(path)
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [_s(c.value) for c in ws[1]]
    if header != _COLS:
        raise ValueError(f"골든 헤더 불일치: {header}")

    ds = DocumentSet(version=golden_version(path))
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = {k: _s(v) for k, v in zip(_COLS, row)}
        if not (r["article_id"] and r["sent_id"]):
            if any(r.values()):
                skipped += 1  # 내용이 있는데 키가 없는 행 — 골든 오손 신호
            continue
        if r["claim_id"]:
            ds.claims.append(ClaimRecord(
                claim_id=r["claim_id"], article_id=r["article_id"], sent_id=r["sent_id"],
                posted_date=r["posted_date"], claim=r["claim"], metric=r["metric"],
                metric_normalized=r["metric_normalized"], value=r["value"], unit=r["unit"],
                value_type=r["value_type"], direction=r["direction"], period=r["period"],
                comparison_basis=r["comparison_basis"],
                comparison_period=r["comparison_period"],
                forecast=(r["forecast"] or "N").upper(),
                kosis_eligible=(r["kosis_eligible"].upper() == "TRUE"),
                exclusion_code=r["exclusion_code"], note=r["note"],
            ))
        else:
            ds.excluded.append(ExcludedRecord(
                article_id=r["article_id"], sent_id=r["sent_id"], sentence=r["claim"],
                exclusion_code=r["exclusion_code"], note=r["note"],
            ))
    if skipped:
        raise ValueError(f"골든 오손 의심: 키(article_id/sent_id) 없는 내용 행 {skipped}개 — 전수 회계 훼손(§8-5)")
    return ds
