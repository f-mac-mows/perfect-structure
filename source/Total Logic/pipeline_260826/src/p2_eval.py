# -*- coding: utf-8 -*-
"""P2 채점기: sentences.jsonl(시스템 분할)을 문장 골든셋과 '경계 위치'로 대조.

두 결과는 같은 정제본의 분할이므로, 공백 제거 공간에서의 경계 위치 집합을 비교한다.
- 경계 재현율: 골든 경계 중 시스템이 맞춘 비율 (놓친 절단)
- 경계 정밀도: 시스템 경계 중 골든에 있는 비율 (과잉 절단)
- 기사 완전 일치: 경계 집합이 완전히 같은 기사 수
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from src import config


def collapse(s: str) -> str:
    return "".join(ch for ch in s if not ch.isspace())


def load_golden(path: Path) -> dict[str, list[str]]:
    wb = load_workbook(path, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = [str(h).strip() if h else "" for h in rows[0]]
    i_id = next(i for i, h in enumerate(header) if "article" in h.lower() and "id" in h.lower())
    i_t = header.index("text")
    out: dict[str, list[str]] = {}
    cur = None
    for r in rows[1:]:
        if r[i_id] is not None and str(r[i_id]).strip():
            cur = str(r[i_id]).strip()
            out[cur] = []
        if r[i_t] is None or not str(r[i_t]).strip():
            continue
        if cur:
            out[cur].append(str(r[i_t]))
    for aid, sents in out.items():
        if len(sents) == 1 and sents[0].strip().lower() == "null":
            out[aid] = []
    return out


def boundary_set(sents: list[str]) -> tuple[set[int], str]:
    """문장 목록 → (경계 위치 집합, 전체 collapse 문자열). 경계 = 각 문장 끝의 누적 위치(마지막 제외)."""
    cum, bounds, parts = 0, set(), []
    for s in sents:
        c = collapse(s)
        parts.append(c)
        cum += len(c)
        bounds.add(cum)
    full = "".join(parts)
    bounds.discard(len(full))
    return bounds, full


def main(argv=None) -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    ap = argparse.ArgumentParser(description="P2 문장 분할을 골든셋과 경계 기준으로 채점")
    ap.add_argument("--system", type=Path, default=Path("data/sentences.jsonl"))
    ap.add_argument("--golden", type=Path,
                    default=config.part1_dir() / "cleaned_sentences_ex.xlsx")
    ap.add_argument("--articles", type=Path, default=Path("data/articles_clean.jsonl"),
                    help="기사 목록 기준 파일 — 0문장(빈 본문) 기사도 대조에 포함하기 위함")
    ap.add_argument("--detail", type=int, default=10, help="불일치 상세 출력 기사 수")
    args = ap.parse_args(argv)

    golden = load_golden(args.golden)
    system: dict[str, list[str]] = defaultdict(list)
    version = "?"
    with open(args.system, encoding="utf-8") as f:
        for line in f:
            s = json.loads(line)
            system[s["article_id"]].append(s["text"])
            version = s.get("pipeline_version", version)
    with open(args.articles, encoding="utf-8") as f:
        all_ids = [json.loads(line)["article_id"] for line in f]
    for aid in all_ids:
        system.setdefault(aid, [])

    tp = fp = fn = 0
    exact = total = 0
    diffs = []
    for aid in all_ids:
        if aid not in golden:
            continue
        # 빈 본문(0문장) 기사도 대조 대상
        total += 1
        gset, gfull = boundary_set(golden[aid])
        sset, sfull = boundary_set(system[aid])
        if gfull != sfull:
            diffs.append((aid, "본문 불일치(보존 오류)", [], []))
            fn += len(gset)
            fp += len(sset)
            continue
        tp += len(gset & sset)
        miss = sorted(gset - sset)
        extra = sorted(sset - gset)
        fn += len(miss)
        fp += len(extra)
        if not miss and not extra:
            exact += 1
        else:
            diffs.append((aid, gfull, miss, extra))

    prec = tp / (tp + fp) if tp + fp else 1.0
    rec = tp / (tp + fn) if tp + fn else 1.0
    f1 = 2 * prec * rec / (prec + rec) if prec + rec else 0.0
    print(f"[{version}] 경계 정밀도 {prec:.1%} · 재현율 {rec:.1%} · F1 {f1:.1%} "
          f"(공통 {tp} · 과잉 {fp} · 누락 {fn})")
    print(f"기사 완전 일치: {exact}/{total}")

    for aid, gfull, miss, extra in diffs[:args.detail]:
        print(f"\n✗ {aid} — 누락 {len(miss)} · 과잉 {len(extra)}")
        for b in miss[:4]:
            print(f"   누락(안 자름): …{gfull[max(0, b - 22):b]} ┃ {gfull[b:b + 22]}…")
        for b in extra[:4]:
            print(f"   과잉(잘못 자름): …{gfull[max(0, b - 22):b]} ┃ {gfull[b:b + 22]}…")
    if len(diffs) > args.detail:
        print(f"\n(그 외 불일치 기사 {len(diffs) - args.detail}건 생략)")


if __name__ == "__main__":
    main()
